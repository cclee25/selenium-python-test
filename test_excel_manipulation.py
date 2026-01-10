from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import os

print("QA Engr B added this print line.")

def download_file():
    driver.find_element(By.ID, "downloadButton").click()

def update_file(file_path, fruit_name, col_name, new_amount):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    get_row = 0
    get_row_1 = 0
    get_column = 0
    get_column_1 = 0

    #Look for price column  
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row = 1, column = i).value == col_name:
            get_column = i
    
    #Look for Fruit Name column  
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row = 1, column = i).value == "fruit_name":
            get_column_1 = i

    #Look for fruit
    for j in range(1, sheet.max_row + 1):
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row = j, column = i).value == fruit_name:
                get_row = j
                break
    
    #Correct Kivi to Kiwi
    for k in range(1, sheet.max_row + 1):
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row = k, column = i).value == "Kivi":
                get_row_1 = k
                break

    sheet.cell(row = get_row_1, column = get_column_1).value = "Kiwi"
    sheet.cell(row = get_row, column = get_column).value = new_amount
    workbook.save(file_path)

#Makes the file path adapt to the OS of the local machine
file_loc = os.path.join(
    os.path.expanduser("~"),
    "Downloads",
    "download.xlsx"
)

fruit = input("Fruit name: ").strip()
updated_amount = input("Fruit price: ").strip()

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

#Get price of a fruit - unique attribute present in an element
#price_column = driver.find_element(By.XPATH, "//div[normalize-space(.)='Price']").get_attribute("data-column-id")
#fruit_price = driver.find_element(By.XPATH, "//div[normalize-space(.)='"+fruit+"']/parent::div"
#"/parent::div/div[@id='cell-"+price_column+"-undefined']")

#Get price of a fruit - no unique attribute in an element                                                                            
headers = driver.find_elements(By.XPATH, "//div[@class='sc-dmyCSP fwQJth rdt_TableHeadRow']/div")
column_index = 0
for header in headers:
    column_index += 1
    if header.find_element(By.XPATH, "div/div").text.strip() == "Price":
        column_index = str(column_index)
        break

fruit_price = driver.find_element(By.XPATH, "//div[@role='row']"
  "[.//div[normalize-space(text())='"+fruit+"']]/div[@role='cell']["+column_index+"]").text

#download excel file
download_file()
time.sleep(7)

#update excel file
update_file(file_loc, fruit, "price", updated_amount)
time.sleep(2)

#upload updated excel file
upload_button = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
upload_button.send_keys(file_loc)

#validate if file is uploaded
wait = WebDriverWait(driver, 5)
upload_notif = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")))
assert upload_notif.text.strip() == "Updated Excel Data Successfully."

#validate if fruit price is updated
fruit_price = driver.find_element(By.XPATH, "//div[@role='row']"
  "[.//div[normalize-space(text())='"+fruit+"']]/div[@role='cell']["+column_index+"]").text
assert fruit_price == updated_amount

time.sleep(10)









